from fastapi import FastAPI, Query, Request
from fastapi_cache.backends.redis import RedisBackend
from fastapi.responses import JSONResponse, FileResponse
from fastapi_cache import FastAPICache
from fastapi_cache.decorator import cache
from datetime import date
import pandas as pd
from typing import Optional, List
import json
import os
from collections.abc import AsyncIterator
from contextlib import asynccontextmanager
from redis import asyncio as aioredis
from openpyxl import Workbook
import tempfile

from app.utils import get_top_100_stocks_on_date
from app.db_handler import db_handler
from app.logger import Logger
from app.utils import (
    fetch_index_performance,
    fetch_index_composition,
    fetch_composition_changes,
)
from app.exceptions import DataNotFoundError

log_base_path = os.path.join(os.getcwd(), "logs")
Logger.setup_logger(log_base_path)


@asynccontextmanager
async def lifespan(app: FastAPI) -> AsyncIterator[None]:
    """
    Lifespan context manager for FastAPI.

    This function handles application startup and shutdown tasks:
    - Initializes Redis-based caching using FastAPI Cache.
    - Sets up the database handler and ensures required tables exist.
    - Cleans any existing cache on startup (optional but useful during development).
    - Gracefully closes the Redis connection on shutdown.
    """
    # Setup Redis
    redis = aioredis.from_url("redis://redis:6379")

    # DB initialize tables
    db_handler.initialize_tables()

    # Setup cache
    FastAPICache.init(RedisBackend(redis), prefix="fastapi-cache")
    await FastAPICache.clear()

    yield

    # Teardown
    await redis.close()


app = FastAPI(title="Custom Equal-Weighted Index API", lifespan=lifespan)


@app.get("/")
async def homepage():
    """
    A home page route
    """
    return {"message": "Welcome to Hedgineer API"}


@app.post("/build-index")
async def build_index(
    start_date: date = Query(...), end_date: Optional[date] = Query(None)
):
    """
    Build the index and calculate performance metrics over a specified date range.

    This endpoint processes index composition and performance data between the provided start and end dates.
    It stores the results in both Redis (for caching) and DuckDB (for persistent storage).

    Args:
        start_date (date): The starting date for the index building process.
        end (date): The ending date for the index building process. Defaults to start_date if not provided.

    Returns a message indicating the success of the operation and the number of days processed.
    """
    end_date = end_date or start_date
    date_range = pd.date_range(start=start_date, end=end_date)
    compositions = []
    performances = []

    previous_index_value = None

    for day in date_range:
        day_str = day.date()
        top_stocks = get_top_100_stocks_on_date(day_str)
        if top_stocks.empty:
            continue

        top_stocks["weight"] = 1 / 100
        top_stocks["date"] = day_str
        top_stocks["notional"] = top_stocks["weight"] * top_stocks["close"]
        index_value = top_stocks["notional"].sum()

        # Calculate daily_return only
        if previous_index_value is None:
            daily_return = 0.0
        else:
            daily_return = (
                ((index_value - previous_index_value) / previous_index_value) * 100
                if previous_index_value != 0
                else 0.0
            )
        previous_index_value = index_value

        performances.append(
            {"date": day_str, "index_value": index_value, "daily_return": daily_return}
        )

        # Cache each day's performance in Redis
        redis_backend = FastAPICache.get_backend()
        perf_key = f"build-index:performance:{day_str}"
        await redis_backend.redis.set(
            perf_key,
            json.dumps({"index_value": index_value, "daily_return": daily_return}),
            ex=86400,
        )

        compositions.append(top_stocks[["date", "ticker", "weight"]])

    if not compositions:
        return JSONResponse(
            status_code=404, content={"message": "No index data could be built."}
        )

    composition_df = pd.concat(compositions)
    db_handler.execute(
        "DELETE FROM index_composition WHERE date BETWEEN ? AND ?",
        (start_date, end_date),
    )
    db_handler.con.register("tmp_comp", composition_df)
    db_handler.execute("INSERT INTO index_composition SELECT * FROM tmp_comp")

    performance_df = pd.DataFrame(performances).sort_values("date")
    db_handler.execute(
        "DELETE FROM index_performance WHERE date BETWEEN ? AND ?",
        (start_date, end_date),
    )
    db_handler.con.register("tmp_perf", performance_df)
    db_handler.execute("INSERT INTO index_performance SELECT * FROM tmp_perf")

    return {
        "message": "Index built and performance calculated.",
        "days_processed": len(performance_df),
    }


@app.get("/index-performance")
@cache(expire=60)
async def index_performance(
    start_date: date = Query(
        ..., description="Start date of the index performance period."
    ),
    end_date: date = Query(
        ..., description="End date of the index performance period."
    ),
):
    """
    Retrieves the index performance (index value, daily return, cumulative return) for the specified date range.
    The function checks Redis cache first, falling back to the database if the data is not cached.

    Args:
        start_date (date): The start date of the performance period.
        end_date (date): The end date of the performance period.

    Returns:
        JSONResponse: A list of dictionaries with index performance data
    """
    Logger.request.info(f"Fetching index performance for {start_date} to {end_date}")
    try:
        redis_backend = FastAPICache.get_backend()
        date_range = pd.date_range(start=start_date, end=end_date)
        results = []
        all_cached = True

        for d in date_range:
            key = f"build-index:performance:{d.date()}"
            cached = await redis_backend.redis.get(key)

            if cached:
                data = json.loads(cached)
                results.append(
                    {
                        "date": str(d.date()),
                        "index_value": data.get("index_value"),
                        "daily_return": data.get("daily_return", 0.0),
                    }
                )
            else:
                all_cached = False
                Logger.signal.info(f"Cache miss for date {d.date()}")
                break

        if not all_cached:
            results = fetch_index_performance(start_date, end_date)

        df_results = pd.DataFrame(results).sort_values("date")
        df_results["cumulative_return"] = df_results["daily_return"].cumsum()

        Logger.response.info(
            f"Index performance returned for {start_date} to {end_date}"
        )
        return df_results.to_dict(orient="records")

    except DataNotFoundError as e:
        return JSONResponse(status_code=404, content={"message": str(e)})

    except Exception as e:
        Logger.response.error(f"Unexpected error: {str(e)}")
        return JSONResponse(
            status_code=500, content={"message": "Internal server error"}
        )


@app.get("/index-composition")
@cache(expire=60)
async def index_composition(date: date = Query(...)):
    """
    Retrieves the index composition for a specific date. Checks Redis cache first;
    if data is not found, falls back to the DuckDB database.

    Args:
        request (Request): FastAPI request object to access app state.
        date (date): The date for which index composition is requested.

    Returns:
        JSONResponse or list[dict]: Composition data or appropriate error message.
    """
    Logger.request.info(f"Fetching index composition for {date}")
    try:
        redis = FastAPICache.get_backend().redis
        key = f"build-index:composition:{date}"

        cached = await redis.get(key)
        if cached:
            tickers = json.loads(cached)
            return [{"ticker": t["ticker"], "weight": t["weight"]} for t in tickers]

        # Fallback to DB
        return fetch_index_composition(date)

    except DataNotFoundError as e:
        Logger.response.warning(str(e))
        return JSONResponse(status_code=404, content={"message": str(e)})

    except Exception as e:
        Logger.response.error(
            f"Unexpected error in /index-composition: {e}", exc_info=True
        )
        return JSONResponse(
            status_code=500,
            content={
                "message": "An unexpected error occurred while fetching index composition."
            },
        )


@app.get("/composition-changes")
@cache(expire=60)
async def composition_changes(
    start_date: date = Query(...), end_date: date = Query(...)
):
    """
    API endpoint to return composition changes (entered/exited tickers) over a date range.

    Args:
        request (Request): FastAPI request object.
        start_date (date): Start date of the range.
        end_date (date): End date of the range.

    Returns:
        JSONResponse: List of changes or error message.
    """
    Logger.request.info(f"Fetching composition changes for {start_date} to {end_date}")
    try:
        changes = fetch_composition_changes(start_date, end_date)

        if not changes:
            Logger.response.info(
                f"No composition changes detected between {start_date} and {end_date}."
            )
            return JSONResponse(
                status_code=200, content={"message": "No composition changes detected."}
            )

        return {"changes": changes}

    except DataNotFoundError as e:
        Logger.response.warning(str(e))
        return JSONResponse(status_code=404, content={"message": str(e)})

    except Exception as e:
        Logger.response.error(
            f"Unexpected error in /composition-changes: {e}", exc_info=True
        )
        return JSONResponse(
            status_code=500,
            content={
                "message": "An unexpected error occurred while processing composition changes."
            },
        )


@app.post("/export-data")
async def export_data(start_date: date = Query(...), end_date: date = Query(...)):
    """
    Exports index performance, composition, and changes between a date range as an Excel file.

    Args:
        start_date (date): Start of export range.
        end_date (date): End of export range.

    Returns:
        FileResponse: Excel file containing exported data.
    """

    def safe_excel_date(d):
        if isinstance(d, pd.Timestamp):
            return d.date()
        elif isinstance(d, tuple) and isinstance(d[0], pd.Timestamp):
            return d[0].date()
        return d

    try:
        performance_list = fetch_index_performance(start_date, end_date)
        df_performance = pd.DataFrame(performance_list).sort_values("date")
        df_performance["cumulative_return"] = df_performance["daily_return"].cumsum()

        wb = Workbook()
        ws = wb.active
        ws.title = "Index Performance"
        ws.append(["Date", "Index Value", "Daily Return", "Cumulative Return"])

        for row in df_performance.itertuples():
            ws.append(
                [
                    safe_excel_date(row.date),
                    row.index_value,
                    row.daily_return,
                    row.cumulative_return,
                ]
            )

        # Index composition
        composition_query = "SELECT date, ticker, weight FROM index_composition WHERE date BETWEEN ? AND ?"
        composition_df = db_handler.fetchdf(composition_query, (start_date, end_date))
        if not composition_df.empty:
            ws_composition = wb.create_sheet(title="Index Composition")
            ws_composition.append(["Date", "Ticker", "Weight"])
            for row in composition_df.itertuples():
                ws_composition.append(
                    [safe_excel_date(row.date), row.ticker, row.weight]
                )

        # Composition changes
        changes = fetch_composition_changes(start_date, end_date)
        if changes:
            ws_changes = wb.create_sheet(title="Composition Changes")
            ws_changes.append(["Date", "Entered Tickers", "Exited Tickers"])
            for change in changes:
                ws_changes.append(
                    [
                        change["date"],
                        ", ".join(change["entered"]),
                        ", ".join(change["exited"]),
                    ]
                )

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        return FileResponse(
            tmp_path,
            filename="index_data.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except DataNotFoundError as e:
        Logger.response.warning(str(e))
        return JSONResponse(status_code=404, content={"message": str(e)})

    except Exception as e:
        Logger.response.error(f"Unexpected error in /export-data: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"message": "An unexpected error occurred while exporting data."},
        )
